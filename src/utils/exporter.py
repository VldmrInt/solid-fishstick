"""
Экспорт данных в различные форматы
"""

import logging
from pathlib import Path
from typing import List
from xml.etree import ElementTree as ET
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from src.parsers.api_parser import ProductInfo

logger = logging.getLogger(__name__)


class DataExporter:
    """Экспорт данных в различные форматы"""

    @staticmethod
    def export_to_excel(products: List[ProductInfo], filename: Path) -> bool:
        """
        Экспортирует товары в Excel с форматированием.

        Args:
            products: Список товаров
            filename: Путь к выходному файлу

        Returns:
            True если успешно, False если ошибка
        """
        if not HAS_OPENPYXL:
            logger.error("openpyxl не установлен, экспорт в Excel невозможен")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Товары"

            # Заголовки
            headers = [
                'SKU',
                'Название',
                'Текущая цена',
                'Старая цена',
                'Рейтинг',
                'Отзывов',
                'Продавец',
                'ИНН',
                'Бренд',
                'Категория',
                'Ссылка',
                'Изображение'
            ]

            ws.append(headers)

            # Стилизация заголовков
            DataExporter._style_header_row(ws)

            # Данные
            for product in products:
                row = [
                    product.sku,
                    product.name,
                    product.current_price,
                    product.original_price,
                    product.rating,
                    product.reviews_count,
                    product.seller_name,
                    product.seller_inn,
                    product.brand,
                    product.category,
                    product.link,
                    product.image_url,
                ]
                ws.append(row)

            # Автоширина столбцов
            DataExporter._auto_width_columns(ws)

            # Сохранение
            wb.save(filename)
            logger.info(f"Excel файл сохранен: {filename}")
            return True

        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            return False

    @staticmethod
    def _style_header_row(ws):
        """Применяет стили к заголовкам"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Высота строки заголовков
        ws.row_dimensions[1].height = 30

    @staticmethod
    def _auto_width_columns(ws):
        """Автоматически подбирает ширину столбцов"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Ограничиваем ширину
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def export_to_xml(products: List[ProductInfo], filename: Path) -> bool:
        """
        Экспортирует товары в XML.

        Args:
            products: Список товаров
            filename: Путь к выходному файлу

        Returns:
            True если успешно, False если ошибка
        """
        try:
            root = ET.Element('products')
            root.set('count', str(len(products)))
            root.set('exported_at', datetime.now().isoformat())

            for product in products:
                item_elem = ET.SubElement(root, 'product')

                # Добавляем все поля
                for field, value in product.to_dict().items():
                    if field not in ['success', 'error']:  # Пропускаем служебные поля
                        elem = ET.SubElement(item_elem, field)
                        elem.text = str(value) if value else ''

            # Форматирование XML
            DataExporter._indent_xml(root)

            # Сохранение
            tree = ET.ElementTree(root)
            tree.write(filename, encoding='utf-8', xml_declaration=True)

            logger.info(f"XML файл сохранен: {filename}")
            return True

        except Exception as e:
            logger.error(f"Ошибка экспорта в XML: {e}")
            return False

    @staticmethod
    def _indent_xml(elem, level=0):
        """Рекурсивно добавляет отступы в XML для красивого форматирования"""
        i = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            for sub_elem in elem:
                DataExporter._indent_xml(sub_elem, level + 1)
            if not sub_elem.tail or not sub_elem.tail.strip():
                sub_elem.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i

    @staticmethod
    def export_to_json(products: List[ProductInfo], filename: Path) -> bool:
        """
        Экспортирует товары в JSON.

        Args:
            products: Список товаров
            filename: Путь к выходному файлу

        Returns:
            True если успешно, False если ошибка
        """
        import json

        try:
            data = {
                'count': len(products),
                'exported_at': datetime.now().isoformat(),
                'products': [p.to_dict() for p in products]
            }

            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            logger.info(f"JSON файл сохранен: {filename}")
            return True

        except Exception as e:
            logger.error(f"Ошибка экспорта в JSON: {e}")
            return False
